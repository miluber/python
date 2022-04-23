from datetime import datetime
from openpyxl import load_workbook

# Copy range of cells as a nested list
# Takes: start cell, end cell, and sheet you want to copy from.


def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    # Loops through selected Rows
    for i in range(startRow, endRow + 1, 1):
        # Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol, endCol+1, 1):
            rowSelected.append(sheet.cell(row=i, column=j).value)
        # Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected

# Paste range
# Paste data from copyRange into template sheet


def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow, endRow+1, 1):
        countCol = 0
        for j in range(startCol, endCol+1, 1):

            sheetReceiving.cell(
                row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1


# this fonction will produce new sheet for a given quater

def createQuaterlySheet(quaterNumber, workbook, sourceSheetName, yearNum):
    newSheetName = 'Iwona' + str(quaterNumber)
    wb.create_sheet(newSheetName)
    sheet = wb[sourceSheetName]
    newsheet = wb[newSheetName]

    if quaterNumber == 1:
        startMonth = 1
        endMonth = 3
    elif quaterNumber == 2:
        startMonth = 4
        endMonth = 6
    elif quaterNumber == 3:
        startMonth = 7
        endMonth = 9

    elif quaterNumber == 4:
        startMonth = 10
        endMonth = 12

    rangeSelected = copyRange(1, 1, 11, 2, sheet)
    pasteRange(1, 1, 11, 2, newsheet, rangeSelected)
    i = 3
    for cell in sheet['A']:
        if isinstance(cell.value, datetime):  # Here we focus ony on cells that have dates
            if cell.value.month >= startMonth and cell.value.month <= endMonth and cell.value.year == yearNum:
                # usre coordinate to get the cell ID can also do row or column
                thisCell = cell.coordinate
                thisRow = cell.row
                thisColumn = cell.column
                print(thisCell + ": " +
                      str(cell.value) + ": " + str(cell.value.month))
                rangeSelected = copyRange(
                    thisColumn, thisRow, 11, thisRow, sheet)
                pasteRange(1, i, 11, i, newsheet, rangeSelected)
                i = i + 1


# Ignore this stuff for now
wb = load_workbook(filename='MyTest.xlsx', data_only=True)
createQuaterlySheet(1, wb, 'Sheet1', 2012)
createQuaterlySheet(2, wb, 'Sheet1', 2012)
createQuaterlySheet(3, wb, 'Sheet1', 2012)
createQuaterlySheet(4, wb, 'Sheet1', 2012)
wb.save('ivona3.xlsx')
