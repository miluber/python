from openpyxl import load_workbook
wb = load_workbook(filename='MyTest.xlsx')
sheet_ranges = wb['2012 Food']
print("Formula: " + sheet_ranges['E10'].value)
wb2 = load_workbook(filename='MyTest.xlsx', data_only=True)
sheet_ranges2 = wb2['2012 Food']
print("Value: " + str(sheet_ranges2['E10'].value))
